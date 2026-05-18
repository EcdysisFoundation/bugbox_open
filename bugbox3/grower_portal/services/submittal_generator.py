from io import BytesIO

from django.db.models import Q
from openpyxl import load_workbook

from bugbox3.core.models import PublicSiteContent

from ..constants import (
    SUBMITTAL_PLANT_TEMPLATE_SLUG,
    SUBMITTAL_PLANT_TYPE,
    SUBMITTAL_SH_END_DEPTH,
    SUBMITTAL_SH_INCHES_END,
    SUBMITTAL_SH_INCHES_START,
    SUBMITTAL_SH_START_DEPTH,
    SUBMITTAL_TEMPLATE_SLUG,
    SUBMITTAL_TEST_ID_PLANT,
    SUBMITTAL_TEST_ID_SH,
    SUBMITTAL_TEST_ID_SOIL,
)
from ..models import GrowerApplication, LabelGeneration


class SubmittalFormGenerator:

    DEPTH_RANGES = [
        (0, 5),
        (5, 10),
        (10, 15),
        (15, 30),
        (30, 60)
    ]

    def __init__(self, cluster_number, year, project_type, label_generation, transect_codes):
        self.cluster_number = cluster_number
        self.year = year
        self.project_type = project_type
        self.label_generation = label_generation
        self.transect_codes = list(transect_codes or [])
        self.forage_transect_codes = []
        self.grower_names = {}
        self._set_forage_codes()

    def _set_forage_codes(self):
        sample_types = self.label_generation.sample_types or [] if self.label_generation else []
        if 'forage' in sample_types:
            self.forage_transect_codes = list(self.transect_codes)
        elif not sample_types and self.transect_codes:
            self.forage_transect_codes = list(self.transect_codes)
        else:
            self.forage_transect_codes = []

    def validate_cluster(self):
        return LabelGeneration.objects.filter(
            cluster_number=self.cluster_number,
            year=self.year,
            project_type=self.project_type,
            label_category='inner',
            status='ready',
        ).exists()

    def get_grower_applications(self):
        """
        Find applications by matching transect codes from label generation.
        """
        if not self.transect_codes:
            return {}

        query = Q()
        for transect_code in self.transect_codes:
            query |= (
                Q(transect_code_1=transect_code) |
                Q(transect_code_2=transect_code) |
                Q(transect_code_3=transect_code) |
                Q(transect_code_4=transect_code)
            )

        applications = GrowerApplication.objects.filter(query).select_related('grower').distinct()

        for app in applications:
            if app.grower:
                grower_name = app.grower.name if hasattr(
                    app.grower, 'name') and app.grower.name else app.grower.username
            else:
                grower_name = app.grower_display_name

            for i in range(1, 5):
                transect_code = getattr(app, f'transect_code_{i}', None)
                if transect_code and transect_code in self.transect_codes:
                    self.grower_names[transect_code] = grower_name

        return self.grower_names

    def _load_template(self, template_slug):
        try:
            template_content = PublicSiteContent.objects.get(title=template_slug)

            if not template_content.file:
                raise ValueError(f"Submittal template '{
                    template_slug}' found in PublicSiteContent but file is missing. Please upload it in Django Admin")

            with template_content.file.open('rb') as file:
                file_content = file.read()
                file_buffer = BytesIO(file_content)
                try:
                    return load_workbook(file_buffer)
                except Exception as e:
                    error_msg = str(e)
                    if 'not a valid' in error_msg.lower() or 'cannot open' in error_msg.lower():
                        raise ValueError(
                            f"Template file '{template_slug}' could not be loaded. "
                            f"Please ensure your template is a valid Excel file (.xlsx). "
                            f"Original error: {error_msg}"
                        ) from e
                    else:
                        raise

        except PublicSiteContent.DoesNotExist:
            raise ValueError(
                f"Submittal template '{template_slug}' not found in PublicSiteContent. "
                f"Please upload the template file to PublicSiteContent with title '{template_slug}' in Django Admin"
            )
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(
                f"Error loading submittal template: {str(e)}"
            ) from e

    def _cm_to_inches(self, cm):
        return round(cm / 2.54)

    def _file_prefix(self):
        return f"{self.project_type}_{self.cluster_number}_{self.year}"

    def _populate_soil_sheet(self, workbook, transect_codes, grower_names):
        try:
            ws = workbook['Soil']
        except KeyError:
            raise ValueError("The Excel template is missing the 'Soil' sheet. Please check the template file.")

        header_row = 6
        headers = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                header_text = str(cell.value).strip().lower()
                headers[header_text] = col_idx

        col_test_id = headers.get('test id')
        col_grower = headers.get('grower')
        col_field = headers.get('field')
        col_sample_id_1 = headers.get('sample id 1')
        col_start_depth = headers.get('start depth')
        col_end_depth = headers.get('end depth')
        col_inches_start = headers.get('inches')
        col_inches_end = col_inches_start + 1 if col_inches_start else None

        required_columns = {
            'test id': col_test_id,
            'grower': col_grower,
            'field': col_field,
            'sample id 1': col_sample_id_1,
            'start depth': col_start_depth,
            'end depth': col_end_depth,
            'inches': col_inches_start,
        }
        missing = [name for name, col in required_columns.items() if col is None]
        if missing:
            raise ValueError(
                f"Error in Soil sheet: Could not find required columns in row 6: {', '.join(missing)}. "
                f"Found columns: {list(headers.keys())}"
            )

        current_row = header_row + 1

        for transect_code in transect_codes:
            grower_name = grower_names.get(transect_code, '')

            for start_depth, end_depth in self.DEPTH_RANGES:
                inches_start = self._cm_to_inches(start_depth)
                inches_end = self._cm_to_inches(end_depth)

                ws.cell(row=current_row, column=col_test_id, value=SUBMITTAL_TEST_ID_SOIL)
                ws.cell(row=current_row, column=col_grower, value=self.cluster_number)
                ws.cell(row=current_row, column=col_field, value=grower_name)
                ws.cell(row=current_row, column=col_sample_id_1, value=transect_code)
                ws.cell(row=current_row, column=col_start_depth, value=start_depth)
                ws.cell(row=current_row, column=col_end_depth, value=end_depth)
                ws.cell(row=current_row, column=col_inches_start, value=inches_start)
                ws.cell(row=current_row, column=col_inches_end, value=inches_end)

                current_row += 1

        return workbook

    def _populate_sh_sheet(self, workbook, transect_codes, grower_names):
        try:
            ws = workbook['SH']
        except KeyError:
            raise ValueError("The Excel template is missing the 'SH' sheet. Please check the template file.")

        header_row = 6
        headers = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                header_text = str(cell.value).strip().lower()
                headers[header_text] = col_idx

        col_test_id = headers.get('test id')
        col_grower = headers.get('grower')
        col_field = headers.get('field')
        col_sample_id_1 = headers.get('sample id 1')
        col_start_depth = headers.get('start depth')
        col_end_depth = headers.get('end depth')
        col_inches_start = headers.get('inches')
        col_inches_end = col_inches_start + 1 if col_inches_start else None

        required_columns = {
            'test id': col_test_id,
            'grower': col_grower,
            'field': col_field,
            'sample id 1': col_sample_id_1,
            'start depth': col_start_depth,
            'end depth': col_end_depth,
            'inches': col_inches_start,
        }
        missing = [name for name, col in required_columns.items() if col is None]
        if missing:
            raise ValueError(
                f"Error in SH sheet: Could not find required columns in row 6: {', '.join(missing)}. "
                f"Found columns: {list(headers.keys())}"
            )

        current_row = header_row + 1

        for transect_code in transect_codes:
            grower_name = grower_names.get(transect_code, '')

            ws.cell(row=current_row, column=col_test_id, value=SUBMITTAL_TEST_ID_SH)
            ws.cell(row=current_row, column=col_grower, value=self.cluster_number)
            ws.cell(row=current_row, column=col_field, value=grower_name)
            ws.cell(row=current_row, column=col_sample_id_1, value=transect_code)
            ws.cell(row=current_row, column=col_start_depth, value=SUBMITTAL_SH_START_DEPTH)
            ws.cell(row=current_row, column=col_end_depth, value=SUBMITTAL_SH_END_DEPTH)
            ws.cell(row=current_row, column=col_inches_start, value=SUBMITTAL_SH_INCHES_START)
            ws.cell(row=current_row, column=col_inches_end, value=SUBMITTAL_SH_INCHES_END)

            current_row += 1

        return workbook

    def _populate_plant_sheet(self, workbook, transect_codes, grower_names):
        """Populate plant submittal form"""
        if not workbook.sheetnames:
            raise ValueError("The Excel template has no sheets. Please check the template file.")

        ws = workbook[workbook.sheetnames[0]]

        header_row = 6
        headers = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                header_text = str(cell.value).strip().lower()
                headers[header_text] = col_idx

        col_test_id = headers.get('test id')
        col_grower = headers.get('grower')
        col_field = headers.get('field')
        col_sample_id_1 = headers.get('sample id 1')
        col_type = headers.get('type')

        required_columns = {
            'test id': col_test_id,
            'grower': col_grower,
            'field': col_field,
            'sample id 1': col_sample_id_1,
            'type': col_type,
        }
        missing = [name for name, col in required_columns.items() if col is None]
        if missing:
            raise ValueError(
                f"Error in Plant sheet: Could not find required columns in row {header_row}: {', '.join(missing)}. "
                f"Found columns: {list(headers.keys())}"
            )

        current_row = header_row + 1

        seen_transects = set()
        for transect_code in transect_codes:
            if transect_code in seen_transects:
                continue
            seen_transects.add(transect_code)

            grower_name = grower_names.get(transect_code, '')

            ws.cell(row=current_row, column=col_test_id, value=SUBMITTAL_TEST_ID_PLANT)
            ws.cell(row=current_row, column=col_grower, value=self.cluster_number)
            ws.cell(row=current_row, column=col_field, value=grower_name)
            ws.cell(row=current_row, column=col_sample_id_1, value=transect_code)
            ws.cell(row=current_row, column=col_type, value=SUBMITTAL_PLANT_TYPE)

            current_row += 1

        return workbook

    def generate_submittal_form(self, generate_soil=True, generate_plant=True):
        if not self.label_generation:
            raise ValueError(
                f"No label generation provided for {self.project_type} cluster "
                f"{self.cluster_number} in year {self.year}"
            )

        if not self.transect_codes:
            raise ValueError('No sample/transect codes selected for submittal generation.')

        grower_names = self.get_grower_applications()

        files = []
        prefix = self._file_prefix()

        if generate_soil:
            workbook = self._load_template(SUBMITTAL_TEMPLATE_SLUG)
            workbook = self._populate_soil_sheet(workbook, self.transect_codes, grower_names)
            workbook = self._populate_sh_sheet(workbook, self.transect_codes, grower_names)

            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            filename = f"Submittal_Form_Soil_{prefix}.xlsx"
            files.append((buffer, filename))

        if generate_plant:
            if not self.forage_transect_codes:
                sample_types = self.label_generation.sample_types or []
                has_forage = 'forage' in sample_types
                if not has_forage:
                    raise ValueError(
                        f"The selected label generation does not include forage sample types. "
                        f"Found sample types: {sample_types}. "
                        "Use a batch with forage labels, or uncheck Plant submittal."
                    )
                raise ValueError(
                    'No codes eligible for plant (forage) submittal in your selection.'
                )

            workbook = self._load_template(SUBMITTAL_PLANT_TEMPLATE_SLUG)
            workbook = self._populate_plant_sheet(workbook, self.forage_transect_codes, grower_names)

            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            filename = f"Submittal_Form_Plant_{prefix}.xlsx"
            files.append((buffer, filename))

        if not files:
            raise ValueError("At least one submittal form type must be selected.")

        return files
